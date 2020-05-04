import json
from flask import Flask, request, render_template, make_response,redirect,url_for
#from form import TestForm
from wtforms import SelectField
from flask_wtf import Form
import boto3
import glob
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
import os
import openpyxl
import datetime
from threading import Timer
import webbrowser


app = Flask(__name__)
app.config['SECRET_KEY'] = "my precious"
fpath = '';
batchname = '';
term = '';
sub = '';
drop1 = '';
accesskey = '';
secretkey = '';
ret='';
class TestForm(Form):
    batch = SelectField(u'', choices=())
    faculty = SelectField(u'', choices=())
    term = SelectField(u'',choices=())
    subject = SelectField(u'',choices=())


@app.route("/", methods=["GET", "POST"])
def index():
    """
    Render form and handle form submission
    """
    form = TestForm(request.form)
    
    form.batch.choices = [('', 'Select a batch')] + [
        (x['Batch_name'], x['Batch_name']) for x in parse_json("data/batch.json")]
    form.faculty.choices = [('','Select Faculty')]+list(set([
        (x['Faculty'],x['Faculty']) for x in parse_json("data/faculty.json")]))
    form.term.choices = [('','Select term')]
    form.subject.choices = [('','Select Subject')]
    if request.method == 'POST':
        global fpath
        fpath = [x['Path'] for x in parse_json("data/batch.json")if x['Batch_name'] == form.batch.data][0]
        
        global accesskey
        accesskey = [x['Access_key'] for x in parse_json("data/batch.json")if x['Batch_name'] == form.batch.data][0]
        
        global secretkey
        secretkey = [x['Secret_key'] for x in parse_json("data/batch.json")if x['Batch_name'] == form.batch.data][0]
        
        global batchname
        batchname = [x['Batch_name'] for x in parse_json("data/batch.json")if x['Batch_name'] == form.batch.data][0]
        
        global term
        term = [x['Term'] for x in parse_json("data/faculty.json")if ((x['Faculty'] == form.faculty.data) and (x['Term'] == form.term.data))][0]
        
        global sub
        sub = [x['Subject'] for x in parse_json("data/faculty.json")if x['Subject'] == form.subject.data][0]
        
        global ret
        ret = "Batch Name: "+batchname+", Term: "+term+", Faculty: "+drop1+", Subject: "+sub
        
        return render_template('frontpage.html',ll=ret)
    return render_template('new.html', form=form)





@app.route("/term/<string:Faculty>/", methods=["GET"])
def get_request(Faculty):
    
    global drop1;
    drop1 = Faculty
    data = list(set([
        ('',x['Term']) for x in parse_json("data/faculty.json")
        if x['Faculty'] == Faculty]))
    response = make_response(json.dumps(data))
    response.content_type = 'application/json'
    return response
    
@app.route("/term1/<string:Term>/", methods=["GET"])
def get_request1(Term):
    
    data = list(set([
        ('', x['Subject']) for x in parse_json("data/faculty.json")
        if (x['Faculty'] == drop1) and (x['Term'] == Term)]))
    response = make_response(json.dumps(data))
    response.content_type = 'application/json'
    return response

@app.route('/newbatch', methods=['POST'])
def newbatch():
    filename = 'data/batch.json'
    
    
    
    with open(filename,"r+") as f:
        new = json.load(f)
        #new.update(data)
        name = request.form['batchname']
        path = request.form['path']
        accesskey = request.form['accesskey']
        secretkey = request.form['secretkey']
           
        flag=0
        if len(new) ==1:
            if ((new[0]['Batch_name']=="Add Batch")*(new[0]['Path']=="Add Path")*(new[0]['Access_key']=="Add Access")*(new[0]['Secret_key']=="Add Secret")):
           
                data = {"Batch_name":name,"Path": path,"Access_key":accesskey,"Secret_key":secretkey}
                new=[data]
                f.seek(0)
                json.dump(new,f,indent=1)
                f.truncate()
                return render_template('new1.html',batch = "First Batch data added")
            
      
        
        for i in range(len(new)):
            if (new[i]['Batch_name']==name):
                new[i]['Path']=path
                new[i]['Access_key']=accesskey
                new[i]['Secret_key']=secretkey
                f.seek(0)
                json.dump(new, f,indent=1)
                f.truncate()
                flag=1
                return render_template('new1.html',batch = "Existing batch details updated or Already existed")
       
            if ((new[i]['Path']==path)*(new[i]['Access_key']==accesskey)&(new[i]['Secret_key']==secretkey)):
                loc=i
                n = new[i]['Batch_name']
                flag=2
                break
            
        if flag==2:
            data = {"Batch_name":name,"Path": path,"Access_key":accesskey,"Secret_key":secretkey}
            new[loc]=data
            f.seek(0)
            json.dump(new, f,indent=1)
            f.truncate()
            return render_template('new1.html',batch = n+"'s name got updated with "+name)
        if flag==0:
            data = {"Batch_name":name,"Path": path,"Access_key":accesskey,"Secret_key":secretkey}
            new = new+[data]
            f.seek(0)
            json.dump(new, f,indent=1)
            return render_template('new1.html',batch = "new batch updated")
            
@app.route('/newfaculty', methods=['POST'])
def newfaculty():
    filename = 'data/faculty.json'                
    
    with open(filename,"r+") as f:
        new = json.load(f)
        facu = request.form['facu']
        ter = request.form['ter']
        subj = request.form['subj']
        chang = request.form['newd']
        fachang = request.form['facd']
        flag=0;
        print("{}-{}".format(chang,fachang))
        if len(new) ==1:
            if ((new[0]['Faculty']=="Add Faculty")*(new[0]['Term']=="Add Term")*(new[0]['Subject']=="Add Subject")):
                chang=subj
                fachang=facu
                data = {"Faculty":fachang,"Term": ter,"Subject":chang}
                new=[data]
                f.seek(0)
                json.dump(new,f,indent=1)
                f.truncate()
                return render_template('new1.html',batch = "First faculty data added")
     
        if (len(chang) == 0 and len(fachang)==0):
            data = {"Faculty":facu,"Term": ter,"Subject":subj}
            print(data)
            if data in new:
                return render_template('new1.html',batch = "Already exists")
            else:
                new = new+[data]
                f.seek(0)
                json.dump(new, f,indent=1)
                return render_template('new1.html',batch = "New faculty data added")
        elif len(fachang)==0:
            for i in range(len(new)):
                if ((new[i]['Faculty']==facu) * (new[i]['Term']==ter)):
                    new[i]['Subject']=chang
                    f.seek(0)
                    json.dump(new, f,indent=1)
                    f.truncate()
                    flag=1;
                    break;
            if flag==1:
                return render_template('new1.html',batch = facu+"'s subject of term "+ter+" is updated with "+chang)
        elif len(chang)==0:
            for i in range(len(new)):
                if ((new[i]['Term']==ter)*(new[i]['Subject']==subj)):
                    new[i]['Faculty']=fachang
                    f.seek(0)
                    json.dump(new, f,indent=1)
                    f.truncate()
                    flag=1
                    break;
            if flag==1:
                return render_template('new1.html',batch = facu+"'s name is updated to "+fachang+" for the term: "+ter+" and subject: "+subj)
        
        else:
            for i in range(len(new)):
                if ((new[i]['Faculty']==facu)*(new[i]['Term']==ter)*(new[i]['Subject']==subj)):
                    new[i]['Faculty']=fachang
                    new[i]['Subject']=chang
                    f.seek(0)
                    json.dump(new, f,indent=1)
                    f.truncate()
                    flag=1
                    break;
            if flag==1:
                return render_template('new1.html',batch = "faculty and Subject name of term: "+ter+" is updated")
        
        if flag==0:
            data = {"Faculty":fachang,"Term": ter,"Subject":chang}
            new = new+[data]
            f.seek(0)
            json.dump(new, f,indent=1)
            return render_template('new1.html',batch = "New faculty added")
            
        

@app.route('/get_results',methods=['POST']) 
def get_results():
    d=request.form['dclass']
    d=datetime.datetime.strptime(d, '%Y-%m-%d')
    d =d.strftime('%d %b %Y')
    img = request.files["image"]

    fname=img.filename
    ext=fname[len(fname)-fname[::-1].find('.')-1:]
    
    exl='\Term '+term+' Attendance'+'.xlsx'

    #fpath=r'F:\e\praxis\Projects\facenet-face-recognition-master\Train Images- AWS'
    imgname=d+ext
    spath=os.path.join(fpath,sub,imgname)
    if not os.path.exists(fpath+'\\'+sub):
        os.mkdir(fpath+'\\'+sub)
    if not os.path.exists(spath):
        img.save(spath)
    #else:
     #   print('Image with that name already exists in folder')
      #  print(keyword)
    if os.path.exists(fpath+exl):
        wb=openpyxl.load_workbook(fpath+exl)
        writer = pd.ExcelWriter(fpath+exl,engine='openpyxl',mode='a')
    else:
        wb=openpyxl.Workbook()
        wb.save(fpath+exl)
        wb=openpyxl.load_workbook(fpath+exl)
        writer = pd.ExcelWriter(fpath+exl,engine='openpyxl',mode='a')
    try:
        ws=wb[sub]
        attendance=pd.read_excel(fpath+exl,sheet_name=sub,keep_default_na=False)
    except:
        wb.create_sheet(sub)
        wb.save(fpath+exl)
        attendance=pd.read_excel(fpath+exl,sheet_name=sub)
        clsdet=pd.read_excel(fpath+'\Class details.xlsx',keep_default_na=False)
        clsdet.to_excel(writer,sheet_name=sub,index=False)
        writer.save()
        writer.close()
        wb=openpyxl.load_workbook(fpath+exl)
        try:
            for i in wb.sheetnames:
                if "Sheet" in i:
                    wb.remove(wb[i])
            wb.save(fpath+exl)
        except:
            wb.save(fpath+exl)
    try:
        attendance=pd.read_excel(fpath+exl,sheet_name=sub,keep_default_na=False)
        attendance['Classes Attended']
        attendance.drop(['Classes Attended','Total Classes','% Attendance'],axis=1,inplace=True)

    except:
        attendance=pd.read_excel(fpath+exl,sheet_name=sub,keep_default_na=False)

    target=fpath+'\\'+sub+'\\'+d+ext
    rekognition = boto3.client("rekognition", 'ap-south-1',aws_access_key_id=accesskey,aws_secret_access_key=secretkey)
    cls=open(target,'rb')
    cls=cls.read()
    clsatt={}
    notdet=[]
    for i in glob.glob(fpath+'\Class Photos\*'):
        img=open(i,'rb')
        res=rekognition.compare_faces(SourceImage={'Bytes': img.read()},TargetImage={'Bytes': cls},SimilarityThreshold=90)
        img.close()
        #print(res['FaceMatches'])
        if len(res['FaceMatches'])==1:
            clsatt[i[len(i)-i[::-1].find('\\'):len(i)-i[::-1].find('.')-1]]=(res['FaceMatches'][0]['Face']['BoundingBox']['Left'],res['FaceMatches'][0]['Face']['BoundingBox']['Top'])
            attendance.loc[attendance['Name']==i[len(i)-i[::-1].find('\\'):len(i)-i[::-1].find('.')-1],d]='P'
        else:
        #   attendance.loc[attendance['Name']==i[len(i)-i[::-1].find('\\'):len(i)-i[::-1].find('.')-1],today]='A'
            notdet.append(i)
    for i in notdet:
        img=open(i,'rb')
        res=rekognition.compare_faces(SourceImage={'Bytes': img.read()},TargetImage={'Bytes': cls},SimilarityThreshold=70)
        img.close()
        if len(res['FaceMatches'])==0:
            attendance.loc[attendance['Name']==i[len(i)-i[::-1].find('\\'):len(i)-i[::-1].find('.')-1],d]='A'
        elif len(res['FaceMatches'])==1:
            loc=(res['FaceMatches'][0]['Face']['BoundingBox']['Left'],res['FaceMatches'][0]['Face']['BoundingBox']['Top'])
            if loc not in clsatt.values():
                attendance.loc[attendance['Name']==i[len(i)-i[::-1].find('\\'):len(i)-i[::-1].find('.')-1],d]='P'
            else:
                attendance.loc[attendance['Name']==i[len(i)-i[::-1].find('\\'):len(i)-i[::-1].find('.')-1],d]='A'
        else:
            attendance.loc[attendance['Name']==i[len(i)-i[::-1].find('\\'):len(i)-i[::-1].find('.')-1],d]='N/A'
    
    dates=list(attendance.columns[2:])
    dates.sort(key=lambda date: datetime.datetime.strptime(date, "%d %b %Y"))
    newcol=list(attendance.columns[0:2])+dates
    attendance=attendance[newcol]

    cols=len(attendance.columns)
    for i in attendance.index:
        attendance.loc[i,'Classes Attended']=sum(attendance.iloc[i,2:]=='P')
    attendance['Total Classes']=cols-2
    attendance['% Attendance']=round(attendance['Classes Attended']*100/attendance['Total Classes'],0)
    writer = pd.ExcelWriter(fpath+exl,engine='openpyxl',mode='a')
    wb=openpyxl.load_workbook(fpath+exl)
    writer.book=wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    attendance.to_excel(writer,sheet_name=sub,index=False)
    writer.save()
    writer.close()
    dash=[]
    wb=openpyxl.load_workbook(fpath+exl)
        
    for k in wb.sheetnames:
        att=pd.read_excel(fpath+exl,sheet_name=k,keep_default_na=False)
        for i in att.index:
            for j in att.columns[2:-3]:
                if att.loc[i,j]=='N/A':
                    dash.append([att.loc[i,'Name'],j,k])
    dashb=pd.DataFrame(dash,columns=['Name','Date','Subject'])
    writer = pd.ExcelWriter(fpath+exl,engine='openpyxl',mode='a')
    writer.book=wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    dashb.to_excel(writer,sheet_name='Dashboard',index=False)
    writer.save()
    writer.close()
        
    wb=openpyxl.load_workbook(fpath+exl)
    for i in wb.sheetnames:
        ws=wb[i]
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = max_length
    wb.save(fpath+exl)
    return render_template('frontpage.html',ll=ret,subj = 'Updated')


    
@app.route("/goback", methods=['POST'])
def goback():
    if request.method == 'POST':
        
        return redirect('/')
    



@app.route("/shutdown", methods = ['POST'])
def shutdown():
    shutdown_server()
    return render_template('new2.html',batch='Server Closed... Please close the window')
    
    



def parse_json(json_file):
    with open(json_file) as data_file:    
        data = json.load(data_file)
    print(data)
    return data


def shutdown_server():
    func = request.environ.get('werkzeug.server.shutdown')
    if func is None:
        raise RuntimeError('Not running with the Werkzeug Server')
    else:
        print('shutdowned')
        #browserExe = "firefox.exe"
        #os.system("taskkill /f /im "+browserExe)
        func()

if __name__ == "__main__":
    
    Timer(2,webbrowser.open('http://127.0.0.1:5000/'))
    app.run(debug=True,use_reloader=False)
    